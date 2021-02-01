import openpyxl

# 데이터 변수 선언
filename = "../Data/sample.xlsx"

# 엑셀 파일 열기
book = openpyxl.load_workbook(filename)

# 시트 추출
sheet = book.worksheets[0]
print("sheet => sheet.title")
print("sheet.min_row => ", sheet.min_row)  # 행(라인)
print("sheet.max_row => ", sheet.max_row)
print("sheet.min_column =>", sheet.min_column)  # 열(칸)
print("sheet.max_column =>", sheet.max_column)

# 데이터 추출
row_data = []
for row in sheet.rows:
    print(row, type(row))

    row_data.append([
        row[0].value,
        row[sheet.max_column - 1].value,
    ])
print('len(row_data) => ', len(row_data), row_data)
