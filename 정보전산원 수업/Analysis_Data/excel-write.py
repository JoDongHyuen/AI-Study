from openpyxl import Workbook

filename = "../Data/sample.xlsx"

wb = Workbook()  # 엑셀 파일 생성, Sheet1 자동 생성
ws = wb.active  # 시트활성화
ws.title = "new sheet"  # 시트명 변경
ws['A1'] = "Language"  # 시트 데이터 삽입
ws['B1'] = "create"

wb.save(filename=filename) # 엑셀파일 저장